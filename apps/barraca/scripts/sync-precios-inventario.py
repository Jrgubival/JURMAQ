#!/usr/bin/env python3
"""
Sincroniza precio / costo / stock de barraca_productos desde el Excel de
inventario del sistema de la barraca.

## Qué toca y qué NO

Toca SOLO: precio, costo, stock.
NO toca (decisión explícita del dueño): nombre, slug, descripcion, imagen,
medida, unidad, categoria_id, destacado, activo, solo_cotizar.

El Excel trae Descripción y Familia, pero los nombres del sitio están
curados a mano y las fotos se asignaron una por una. Pisarlos con el texto
del sistema de inventario ("ABRAZ. FIJACION TUBOS C/CO...") destruiría meses
de trabajo y arruinaría el SEO. Por eso este script ignora esas columnas.

## Reglas de seguridad de los datos

1. **Nunca escribe precio 0.** El Excel trae 8 productos en 0. La tabla
   barraca_precio_historial tiene CHECK (precio > 0), así que un 0 rompería
   el trigger de historial; y un producto a $0 en un ecommerce es una venta
   regalada. Se saltan y se reportan.

1b. **Nunca escribe un precio bajo el costo.** El propio Excel trae 3 filas
   con precio <= costo (ej. TAPA GORRO PPR SO 20: costo 64, precio 50). Es un
   error del sistema de inventario, no una oferta. Se saltan.

1c. **Retiene las bajas mayores a 50% para revisión humana.** Son casi
   siempre un desajuste de ENVASE, no un precio nuevo. Caso real detectado:
   "Tarugo Nylon N#10 Bolsa 500 UN" figuraba en el sitio a $24.020 y el Excel
   traía $50 — porque la fila del Excel es por UNIDAD y el sitio vende la
   bolsa de 500 (500 x 50 = 25.000, que cuadra con el precio del sitio).
   Aplicarlo habría vendido 500 tarugos a $50.

   El umbral es asimétrico a propósito: una baja equivocada se vende bajo
   costo y la plata no vuelve; un alza equivocada solo hace que ese producto
   no se venda, y se corrige cuando alguien lo note. Por eso las alzas SÍ se
   aplican (varias son correcciones necesarias: el "Juego Llaves Hexagonales"
   estaba a $4.160 con costo $13.012, o sea vendiéndose a pérdida).

   Las retenidas quedan en scripts/backups/revisar-precios-<fecha>.csv.

2. **Respeta ofertas activas.** Si un producto tiene en_oferta = true, su
   campo `precio` es el precio CON descuento y `precio_original` el de lista.
   Escribir el precio de lista encima borraría la oferta sin avisar. En esos
   casos se actualiza `precio_original` (que es lo que el Excel realmente
   representa) y se deja `precio` intacto.

3. **Stock negativo se sube a 0.** El sistema de inventario arrastra stocks
   negativos (hasta -225) por descuadres. Publicar stock negativo rompe la
   lógica de disponibilidad del sitio.

4. **Dry-run por defecto.** Sin --apply no escribe nada. Con --apply guarda
   antes un backup JSON del estado actual en scripts/backups/.

## Uso

    # ver qué pasaría (no escribe)
    python3 sync-precios-inventario.py "/ruta/INVENTARIO.xlsx"

    # aplicar de verdad
    python3 sync-precios-inventario.py "/ruta/INVENTARIO.xlsx" --apply

Necesita openpyxl. Si no está: python3 -m venv .venv && .venv/bin/pip install openpyxl
Lee credenciales de apps/barraca/.env.local (NEXT_PUBLIC_SUPABASE_URL y
SUPABASE_SERVICE_ROLE_KEY). Nunca las imprime.
"""

import json
import os
import sys
import time
import urllib.error
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

REPO = Path(__file__).resolve().parents[3]
ENV_FILE = REPO / "apps" / "barraca" / ".env.local"
BACKUP_DIR = Path(__file__).resolve().parent / "backups"

# Columnas del Excel que consumimos. El resto se ignora a propósito.
COL_CODIGO = "Código"
COL_PRECIO = "Precio General"
COL_COSTO = "Costo Bru."
COL_STOCK = "Stock Total"


def cargar_env() -> tuple[str, str]:
    """Lee url + service_role de .env.local sin imprimirlos nunca."""
    if not ENV_FILE.exists():
        sys.exit(f"No existe {ENV_FILE}")
    env = {}
    for linea in ENV_FILE.read_text(encoding="utf-8").splitlines():
        linea = linea.strip()
        if not linea or linea.startswith("#") or "=" not in linea:
            continue
        k, v = linea.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    url = env.get("NEXT_PUBLIC_SUPABASE_URL")
    key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        sys.exit("Faltan NEXT_PUBLIC_SUPABASE_URL o SUPABASE_SERVICE_ROLE_KEY en .env.local")
    return url.rstrip("/"), key


def rest(url: str, key: str, path: str, method: str = "GET", body=None, extra_headers=None):
    """Llamada a PostgREST. Reintenta una vez ante 5xx / error de red."""
    req_url = f"{url}/rest/v1/{path}"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if extra_headers:
        headers.update(extra_headers)
    data = json.dumps(body).encode("utf-8") if body is not None else None

    for intento in (1, 2):
        req = urllib.request.Request(req_url, data=data, headers=headers, method=method)
        try:
            with urllib.request.urlopen(req, timeout=60) as resp:
                raw = resp.read().decode("utf-8")
                return json.loads(raw) if raw.strip() else []
        except urllib.error.HTTPError as e:
            detalle = e.read().decode("utf-8", "replace")[:400]
            if 500 <= e.code < 600 and intento == 1:
                time.sleep(2)
                continue
            # El detalle de PostgREST no lleva secretos; la URL sí lleva el
            # proyecto pero no la key (va en headers).
            raise SystemExit(f"HTTP {e.code} en {method} {path}: {detalle}")
        except urllib.error.URLError as e:
            if intento == 1:
                time.sleep(2)
                continue
            raise SystemExit(f"Red caída en {method} {path}: {e.reason}")
    return []


def leer_excel(ruta: str) -> dict[str, dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("Falta openpyxl. Instálalo en un venv: python3 -m venv .venv && .venv/bin/pip install openpyxl")

    wb = load_workbook(ruta, data_only=True, read_only=True)
    ws = wb["Hoja2"] if "Hoja2" in wb.sheetnames else wb[wb.sheetnames[0]]

    filas = ws.iter_rows(values_only=True)
    encabezados = [str(c).strip() if c is not None else "" for c in next(filas)]
    idx = {h: i for i, h in enumerate(encabezados)}
    for col in (COL_CODIGO, COL_PRECIO, COL_COSTO, COL_STOCK):
        if col not in idx:
            sys.exit(f"El Excel no tiene la columna {col!r}. Tiene: {encabezados}")

    def num(v):
        if v is None:
            return None
        try:
            return int(round(float(v)))
        except (TypeError, ValueError):
            return None

    items: dict[str, dict] = {}
    for fila in filas:
        if not fila:
            continue
        codigo = fila[idx[COL_CODIGO]]
        if codigo is None or str(codigo).strip() == "":
            continue
        codigo = str(codigo).strip()
        items[codigo] = {
            "precio": num(fila[idx[COL_PRECIO]]),
            "costo": num(fila[idx[COL_COSTO]]),
            "stock": num(fila[idx[COL_STOCK]]),
        }
    wb.close()
    return items


def traer_productos(url: str, key: str) -> list[dict]:
    """Trae todos los productos paginando (PostgREST corta en 1000)."""
    campos = "id,codigo,nombre,precio,costo,stock,en_oferta,precio_original,activo"
    out, offset, page = [], 0, 1000
    while True:
        lote = rest(
            url, key,
            f"barraca_productos?select={campos}&order=id.asc&limit={page}&offset={offset}",
        )
        out.extend(lote)
        if len(lote) < page:
            break
        offset += page
    return out


def main() -> None:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    aplicar = "--apply" in sys.argv
    if not args:
        sys.exit(__doc__)
    ruta_excel = args[0]

    url, key = cargar_env()
    excel = leer_excel(ruta_excel)
    productos = traer_productos(url, key)

    print(f"Excel: {len(excel)} códigos · BD: {len(productos)} productos")

    por_codigo = {}
    sin_codigo = 0
    for p in productos:
        c = (p.get("codigo") or "").strip()
        if not c:
            sin_codigo += 1
            continue
        por_codigo[c] = p

    cambios, saltados, sin_match_bd, retenidos = [], [], [], []
    for codigo, p in por_codigo.items():
        fila = excel.get(codigo)
        if fila is None:
            sin_match_bd.append(p)
            continue

        nuevo_precio = fila["precio"]
        if not nuevo_precio or nuevo_precio <= 0:
            saltados.append((p, "precio 0 o vacío en el Excel"))
            continue

        # Guarda 1: el propio Excel a veces trae precio <= costo. Es un error
        # de su sistema, no una liquidación. Nunca publicamos a pérdida.
        costo_excel = fila["costo"]
        if costo_excel and nuevo_precio <= costo_excel:
            saltados.append((p, f"precio {nuevo_precio:,} <= costo {costo_excel:,} en el Excel"))
            continue

        # Guarda 2: baja mayor a 50% = casi siempre desajuste de envase
        # (fila por unidad vs producto vendido por bolsa/paquete). Se retiene
        # para revisión en vez de aplicarse; ver docstring.
        precio_actual = p.get("precio")
        if not p.get("en_oferta") and precio_actual and nuevo_precio < precio_actual * 0.5:
            caida = (nuevo_precio - precio_actual) / precio_actual * 100
            razon = "posible desajuste de envase (unidad vs paquete)"
            if costo_excel:
                razon += f"; costo unitario Excel ${costo_excel:,}"
            retenidos.append((p, nuevo_precio, caida, razon))
            continue

        upd: dict = {}
        # Oferta activa: el Excel trae el precio de LISTA, que corresponde a
        # precio_original. Tocar `precio` borraría el descuento vigente.
        if p.get("en_oferta"):
            if p.get("precio_original") != nuevo_precio:
                upd["precio_original"] = nuevo_precio
        else:
            if p.get("precio") != nuevo_precio:
                upd["precio"] = nuevo_precio

        costo = fila["costo"]
        if costo is not None and costo > 0 and p.get("costo") != costo:
            upd["costo"] = costo

        stock = fila["stock"]
        if stock is not None:
            stock = max(0, stock)  # el inventario arrastra negativos
            if p.get("stock") != stock:
                upd["stock"] = stock

        if upd:
            cambios.append((p, upd))

    solo_en_excel = [c for c in excel if c not in por_codigo]

    # --- Informe ---------------------------------------------------------
    print()
    print(f"  Con cambios         : {len(cambios)}")
    print(f"  Sin cambios         : {len(por_codigo) - len(cambios) - len(saltados) - len(retenidos) - len(sin_match_bd)}")
    print(f"  Saltados (0 o bajo costo)   : {len(saltados)}")
    print(f"  RETENIDOS para revisión     : {len(retenidos)}")
    print(f"  En BD sin fila Excel: {len(sin_match_bd)}")
    print(f"  En Excel sin producto en BD: {len(solo_en_excel)}")
    print(f"  Productos en BD sin código  : {sin_codigo}")

    subidas = [(p, u) for p, u in cambios if "precio" in u and p["precio"] and u["precio"] > p["precio"]]
    bajadas = [(p, u) for p, u in cambios if "precio" in u and p["precio"] and u["precio"] < p["precio"]]
    print()
    print(f"  Precios que SUBEN : {len(subidas)}")
    print(f"  Precios que BAJAN : {len(bajadas)}")

    def variacion(par):
        p, u = par
        return (u["precio"] - p["precio"]) / p["precio"] if p.get("precio") else 0

    for etiqueta, grupo in (("MAYORES ALZAS", subidas), ("MAYORES BAJAS", bajadas)):
        if not grupo:
            continue
        print(f"\n  {etiqueta} (top 8):")
        for p, u in sorted(grupo, key=variacion, reverse=(etiqueta == "MAYORES ALZAS"))[:8]:
            pct = variacion((p, u)) * 100
            print(f"    {pct:+7.1f}%  {p['precio']:>9,} → {u['precio']:>9,}  {p['nombre'][:46]}")

    if saltados:
        print(f"\n  SALTADOS ({len(saltados)}):")
        for p, motivo in saltados[:12]:
            print(f"    [{p['codigo']}] {p['nombre'][:46]} — {motivo}")

    if retenidos:
        print(f"\n  RETENIDOS — bajas >50%, revisar a mano ({len(retenidos)}):")
        for p, nuevo, caida, razon in sorted(retenidos, key=lambda r: r[2]):
            print(f"    {caida:+7.1f}%  {p['precio']:>9,} → {nuevo:>8,}  [{p['codigo']}] {p['nombre'][:40]}")
            print(f"             {razon}")

    ofertas = [(p, u) for p, u in cambios if p.get("en_oferta")]
    if ofertas:
        print(f"\n  EN OFERTA — se actualiza precio_original, NO el precio con descuento ({len(ofertas)}):")
        for p, u in ofertas[:10]:
            print(f"    [{p['codigo']}] {p['nombre'][:44]} · oferta ${p['precio']:,} · lista → ${u.get('precio_original', p.get('precio_original')):,}")

    if not aplicar:
        print("\n=== DRY-RUN. No se escribió nada. Repite con --apply para aplicar. ===")
        return

    # --- Aplicar ---------------------------------------------------------
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    sello = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    backup = BACKUP_DIR / f"precios-antes-{sello}.json"
    backup.write_text(
        json.dumps(
            [
                {k: p[k] for k in ("id", "codigo", "precio", "costo", "stock", "en_oferta", "precio_original")}
                for p in productos
            ],
            ensure_ascii=False,
            indent=1,
        ),
        encoding="utf-8",
    )
    print(f"\nBackup del estado anterior: {backup}")

    # Los retenidos se dejan en CSV para que el dueño los resuelva a mano:
    # cada uno necesita decidir si el producto del sitio es un paquete (y hay
    # que multiplicar el precio unitario) o si el precio del sitio estaba mal.
    if retenidos:
        csv_path = BACKUP_DIR / f"revisar-precios-{sello}.csv"
        filas = ["codigo,nombre,precio_sitio_actual,precio_excel,variacion_pct,motivo"]
        for p, nuevo, caida, razon in sorted(retenidos, key=lambda r: r[2]):
            nombre = '"' + str(p["nombre"]).replace('"', "'") + '"'
            filas.append(f'{p["codigo"]},{nombre},{p["precio"]},{nuevo},{caida:.1f},"{razon}"')
        csv_path.write_text("\n".join(filas) + "\n", encoding="utf-8")
        print(f"Retenidos para revisar a mano: {csv_path}")

    ok = fallos = 0
    for i, (p, upd) in enumerate(cambios, 1):
        try:
            rest(
                url, key,
                f"barraca_productos?id=eq.{p['id']}",
                method="PATCH",
                body=upd,
                extra_headers={"Prefer": "return=minimal"},
            )
            ok += 1
        except SystemExit as e:
            fallos += 1
            print(f"  FALLO id={p['id']} [{p['codigo']}]: {e}")
        if i % 200 == 0:
            print(f"  {i}/{len(cambios)}…")

    print(f"\nListo: {ok} actualizados, {fallos} fallidos.")
    print(f"Para revertir: los valores previos están en {backup}")


if __name__ == "__main__":
    main()
